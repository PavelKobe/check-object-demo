"""
export_excel.py — generate Excel file for checks list
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_checks_excel(checks: list[dict]) -> bytes:
    """Generate Excel bytes for the checks list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Проверки"

    headers = ["#", "Дата / Время", "Магазин", "По штату", "По факту", "Балл", "Категория", "Коммент."]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    for row_idx, (c, i) in enumerate(zip(checks, range(len(checks), 0, -1)), 2):
        p = c.get("payload") or {}
        total_staff = p.get("b2_total_staff")
        b2_checks = p.get("b2_checks") or []
        if b2_checks:
            ap = b2_checks[0].get("absent_posts")
        else:
            ap = p.get("b2_absent_posts")
        fact = (total_staff - ap) if isinstance(total_staff, (int, float)) and isinstance(ap, (int, float)) else "—"
        comments_count = len(p.get("comments") or {})

        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=c.get("timestamp", ""))
        ws.cell(row=row_idx, column=3, value=c.get("store_name", ""))
        ws.cell(row=row_idx, column=4, value=total_staff if total_staff is not None else "—")
        ws.cell(row=row_idx, column=5, value=fact)
        ws.cell(row=row_idx, column=6, value=f"{c.get('total_score', '')}%")
        ws.cell(row=row_idx, column=7, value=c.get("grade", ""))
        ws.cell(row=row_idx, column=8, value=comments_count if comments_count else "—")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(headers[col - 1]) + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
